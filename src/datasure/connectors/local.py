import os
from pathlib import Path

import polars as pl
import streamlit as st
from openpyxl import load_workbook

from datasure.utils import duckdb_get_table, duckdb_save_table
from datasure.utils.file_security import (
    SecurityError,
    VirusScanner,
    get_file_info,
    secure_read_data,
    validate_file_security,
)

# --- Get List of sheet from excel ---#


def local_excel_sheet_names(file_path: str) -> list:
    """Import an excel file and return the list of sheet names.

    PARAMS:
    -------
    file_path: str : path to the excel file
    """
    excel_file = load_workbook(file_path, read_only=True)
    sheet_names = excel_file.sheetnames

    return sheet_names


def local_read_data(filename: str, sheet_name: str | None = None) -> pl.DataFrame:
    """Import data from a file with enhanced security validation.

    PARAMS:
    -------
    filename: str : path to the file
    sheet_name: str : name of the sheet to import (only for excel files)

    Returns
    -------
    data: pl.DataFrame : imported data

    Raises
    ------
    SecurityError: If file fails security validation
    """
    try:
        # Use secure reading function with comprehensive validation
        return secure_read_data(filename, sheet_name)
    except SecurityError:
        # Re-raise security errors for UI handling
        raise
    except Exception as e:
        # Convert other errors to security errors for consistent handling
        raise SecurityError(f"File reading failed: {e!s}") from e


# --- FORM for Adding file from local storage ---#


def local_add_form(
    project_id: str, edit_mode: bool = False, defaults: dict | None = None
) -> None:
    """Form for adding a file from local storage.

    PARAMS:
    -------
    None

    Returns
    -------
    None

    """
    mode = "edit" if edit_mode else "add"

    def valid_alias(alias: str) -> bool:
        """Validate alias for uniqueness and length."""
        if not alias:
            st.error("Alias cannot be empty")
            return False
        if len(alias) > 20:
            st.error("Alias must be a maximum of 20 characters")
            return False
        return True

    def enhanced_valid_file_path(file_path: str) -> tuple[bool, str]:
        """Enhanced file path validation with comprehensive security checks.

        Returns
        -------
            Tuple[bool, str]: (is_valid, message)
        """
        if not file_path:
            return False, "File path cannot be empty"

        if not os.path.isfile(file_path):
            return False, "File not found. Please check the file path"

        # Enhanced security validation
        is_valid, error_msg = validate_file_security(file_path)
        if not is_valid:
            return False, error_msg

        return True, "File validation successful"

    # Get the path to the assets directory relative to the package
    assets_dir = Path(__file__).parent.parent / "assets"
    image_path = assets_dir / "hard-disk.png"
    st.image(str(image_path), width=100)
    st.subheader("Add File from Local Storage")

    # Security status indicator
    with st.expander(":material/security: File Security Status", expanded=False):
        st.markdown(":material/check: File size limits enforced")
        st.markdown(":material/check: Content type validation enabled")
        st.markdown(":material/check: Malicious content detection active")

        # Virus scanning status
        virus_scan_available = VirusScanner.is_available()
        if virus_scan_available:
            st.markdown(":material/verified: Virus scanning available")
        else:
            st.markdown(":material/info: Virus scanning not available")

        st.markdown("**Maximum file sizes:**")
        st.markdown("- CSV: 100MB")
        st.markdown("- Excel: 50MB")
        st.markdown("- JSON: 10MB")
        st.markdown("- Stata: 100MB")

    if edit_mode:
        st.info("You are in edit mode. Please modify the file details below.")
        # load the current file details from the defaults
        default_local_file_alias = defaults.get("alias", "")
        default_local_added_file = defaults.get("filename", "")
        default_local_added_file_sheet_name = defaults.get("sheet_name", "")
    else:
        # set default values for the form inputs
        default_local_file_alias = ""
        default_local_added_file = ""
        default_local_added_file_sheet_name = ""

    local_file_alias = st.text_input(
        label="alias*",
        help="Enter a unique, short, descriptive name for the file",
        placeholder=default_local_file_alias if edit_mode else "",
        disabled=edit_mode,
    )
    if local_file_alias:
        valid_alias(local_file_alias)

    # file uploader. Limit to 1 file and allow only file types selected
    local_added_file = st.text_input(
        label="file path*",
        help="Add full file name and path. eg. C:/data/survey.dta",
        placeholder=default_local_added_file if edit_mode else "",
    )

    if local_added_file:
        # Enhanced security validation with detailed feedback
        is_valid, validation_msg = enhanced_valid_file_path(local_added_file)

        if not is_valid:
            st.error(f":material/error: {validation_msg}")
            local_added_file_sheet_name = ""
        else:
            # Show file passed validation
            st.success(f":material/check_circle: {validation_msg}")

            # Display file information
            file_info = get_file_info(local_added_file)

            with st.expander(":material/info: File Information", expanded=False):
                col1, col2 = st.columns(2)

                with col1:
                    st.markdown(f"**Size:** {file_info['size_mb']} MB")
                    st.markdown(f"**Type:** {file_info['extension']}")

                with col2:
                    st.markdown(f"**Hash:** `{file_info['hash_short']}...`")

                    # Optional virus scan
                    if VirusScanner.is_available() and st.button(
                        ":material/scan_virus: Scan for viruses",
                        key=f"virus_scan_{hash(local_added_file)}",
                    ):
                        with st.spinner("Scanning file..."):
                            is_clean, scan_result = VirusScanner.scan_file(
                                local_added_file
                            )
                            if is_clean:
                                st.success(f":material/verified: {scan_result}")
                            else:
                                st.error(f":material/dangerous: {scan_result}")

            # Handle Excel sheet selection
            local_added_file_ext = local_added_file.split(".")[-1]
            if local_added_file_ext in ["xlsx", "xls"]:
                try:
                    sheets = local_excel_sheet_names(local_added_file)

                    # check if default sheetname exists in the list of sheets
                    if (
                        default_local_added_file_sheet_name
                        and default_local_added_file_sheet_name in sheets
                    ):
                        # get index of the default sheet name or set to first sheet
                        default_sheet_index = sheets.index(
                            default_local_added_file_sheet_name
                        )
                    else:
                        default_sheet_index = 0

                    local_added_file_sheet_name = st.selectbox(
                        label="Sheet Name",
                        options=sheets,
                        index=default_sheet_index,
                    )
                except Exception as e:
                    st.warning(f":material/warning: Could not read Excel sheets: {e!s}")
                    local_added_file_sheet_name = ""
            else:
                local_added_file_sheet_name = ""
    else:
        local_added_file_sheet_name = ""

    # add a submit button
    local_add_btn = st.button(
        "Add File",
        type="primary",
        use_container_width=True,
        key=f"add_file_key{mode}",
        disabled=not local_added_file and not local_file_alias,
    )

    st.markdown("**required*")

    # if submit (local_add_file) button is clicked
    if local_add_btn:
        import_log = duckdb_get_table(project_id, alias="import_log", db_name="logs")

        # check that alias is unique
        if not import_log.is_empty() and (
            local_file_alias in import_log["alias"].to_list()
        ):
            st.error(
                "Alias already exists. Please choose a different alias or edit the existing one."
            )
        else:
            if edit_mode:
                # update the row in the cache file
                import_log = import_log.with_columns(
                    pl.when(pl.col("alias") == default_local_file_alias)
                    .then(pl.lit(local_added_file))
                    .otherwise(pl.col("filename"))
                    .alias("filename"),
                )
        import_log = duckdb_get_table(project_id, alias="import_log", db_name="logs")

        # check that alias is unique
        if not import_log.is_empty() and (
            local_file_alias in import_log["alias"].to_list()
        ):
            st.error(
                "Alias already exists. Please choose a different alias or edit the existing one."
            )
        else:
            if edit_mode:
                # update the row in the cache file
                import_log = import_log.with_columns(
                    pl.when(pl.col("alias") == default_local_file_alias)
                    .then(pl.lit(local_added_file))
                    .otherwise(pl.col("filename"))
                    .alias("filename"),
                )

            else:
                # create a new row with the file details
                new_row = {
                    "refresh": True,
                    "load": True,
                    "source": "local storage",
                    "alias": local_file_alias,
                    "filename": local_added_file,
                    "sheet_name": local_added_file_sheet_name,
                    "server": "",
                    "form_id": "",
                    "private_key": "",
                    "save_to": "",
                    "attachments": False,
                }

                # append the new row to the cache file
                import_log = pl.concat(
                    [import_log, pl.DataFrame([new_row])], how="vertical"
                )

            # save the updated cache file
            duckdb_save_table(
                project_id,
                import_log,
                alias="import_log",
                db_name="logs",
            )


# --- Load data from local storage ---#


def local_load_action(
    project_id: str, alias: str, filename: str, sheet_name: str | None
) -> None:
    """Load data from local storage with enhanced security validation.

    PARAMS:
    -------
    project_id: str : project ID
    alias: str : alias for the data
    filename: str : path to the file
    sheet_name: str : name of the sheet to import (if applicable)

    Returns
    -------
    None

    Raises
    ------
    SecurityError: If file fails security validation
    """
    try:
        # Use enhanced secure data reading
        with st.spinner(f"Loading {filename} with security validation..."):
            data: pl.DataFrame = local_read_data(filename, sheet_name)

        # Log successful load with file info
        file_info = get_file_info(filename)
        st.success(
            f":material/check_circle: Successfully loaded {data.height:,} rows "
            f"and {data.width} columns from {file_info['extension']} file"
        )

        # save data to DuckDB
        duckdb_save_table(
            project_id,
            data,
            alias=alias,
            db_name="raw",
        )

    except SecurityError as e:
        st.error(f":material/security: Security validation failed: {e!s}")
        raise
    except Exception as e:
        st.error(f":material/error: Failed to load file: {e!s}")
        raise
