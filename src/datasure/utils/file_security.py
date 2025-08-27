"""Enhanced file upload security utilities for DataSure.

This module provides comprehensive file validation, content scanning,
and security checks to prevent malicious file uploads and DoS attacks.
"""

import hashlib
import json
import logging
import os
import re
import subprocess
from pathlib import Path
from typing import Any, ClassVar

import pandas as pd
import polars as pl

logger = logging.getLogger(__name__)


class FileSecurityConfig:
    """Configuration constants for file upload security."""

    # Maximum file sizes (in bytes)
    MAX_FILE_SIZES: ClassVar[dict[str, int]] = {
        "csv": 100 * 1024 * 1024,  # 100MB
        "xlsx": 50 * 1024 * 1024,  # 50MB
        "xls": 50 * 1024 * 1024,  # 50MB
        "json": 10 * 1024 * 1024,  # 10MB
        "dta": 100 * 1024 * 1024,  # 100MB
    }

    # MIME type validation mapping
    ALLOWED_MIME_TYPES: ClassVar[dict[str, list[str]]] = {
        "csv": ["text/csv", "text/plain", "application/csv"],
        "xlsx": ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"],
        "xls": ["application/vnd.ms-excel"],
        "json": ["application/json", "text/json"],
        "dta": ["application/x-stata-dta", "application/octet-stream"],
    }

    # Content validation limits
    MAX_ROWS_PREVIEW = 1000
    MAX_COLUMNS = 1000
    MAX_CONCURRENT_UPLOADS = 3

    # Suspicious content patterns
    SUSPICIOUS_PATTERNS: ClassVar[list[str]] = [
        r"<script[^>]*>.*?</script>",  # JavaScript
        r"javascript:",  # JavaScript URLs
        r"data:.*base64",  # Data URLs
        r"<\?php.*\?>",  # PHP code
        r"<%.*%>",  # ASP/JSP code
        r"=\s*cmd\s*\|",  # Command injection
        r"union\s+select",  # SQL injection
        r"<iframe[^>]*>",  # Iframe injection
        r"<object[^>]*>",  # Object embedding
        r"<embed[^>]*>",  # Embed tags
    ]


class SecurityError(Exception):
    """Custom exception for security-related file validation errors."""


def validate_file_security(file_path: str) -> tuple[bool, str | None]:
    """Comprehensive file security validation.

    Args:
        file_path: Path to the file to validate

    Returns
    -------
        Tuple[bool, Optional[str]]: (is_valid, error_message)
    """
    try:
        file_path_obj = Path(file_path)

        # 1. File existence and accessibility
        if not file_path_obj.exists():
            return False, "File does not exist"

        if not file_path_obj.is_file():
            return False, "Path is not a file"

        # 2. File extension validation
        file_ext = file_path_obj.suffix.lower().lstrip(".")
        if file_ext not in FileSecurityConfig.MAX_FILE_SIZES:
            return False, f"Unsupported file type: {file_ext}"

        # 3. File size validation
        file_size = file_path_obj.stat().st_size
        max_size = FileSecurityConfig.MAX_FILE_SIZES[file_ext]
        if file_size > max_size:
            max_size_mb = max_size // (1024 * 1024)
            return (
                False,
                f"File too large. Maximum size for {file_ext}: {max_size_mb}MB",
            )

        # 4. MIME type validation
        if not _validate_mime_type(file_path, file_ext):
            return False, "File content doesn't match expected format"

        # 5. File content validation
        if not _validate_file_content(file_path, file_ext):
            return False, "File content validation failed"

    except Exception as e:
        logger.exception(f"Security validation error for {file_path}")
        return False, f"Security validation error: {e!s}"
    else:
        return True, None


def _validate_mime_type(file_path: str, expected_ext: str) -> bool:
    """Validate file MIME type matches extension.

    Args:
        file_path: Path to the file
        expected_ext: Expected file extension

    Returns
    -------
        bool: True if MIME type is valid
    """
    try:
        # Try to use python-magic if available
        try:
            import magic

            mime_type = magic.from_file(file_path, mime=True)
            allowed_types = FileSecurityConfig.ALLOWED_MIME_TYPES.get(expected_ext, [])
        except ImportError:
            # Fallback to basic validation if magic is not available
            logger.warning("python-magic not available, skipping MIME validation")
            return True
        else:
            return mime_type in allowed_types
    except Exception as e:
        logger.warning(f"MIME type validation failed for {file_path}: {e}")
        return True  # Don't block if MIME detection fails


def _validate_file_content(file_path: str, file_ext: str) -> bool:
    """Basic content validation to detect malicious files.

    Args:
        file_path: Path to the file
        file_ext: File extension

    Returns
    -------
        bool: True if content is valid
    """
    try:
        if file_ext == "csv":
            return _validate_csv_content(file_path)
        elif file_ext in ["xlsx", "xls"]:
            return _validate_excel_content(file_path)
        elif file_ext == "json":
            return _validate_json_content(file_path)
        elif file_ext == "dta":
            return _validate_stata_content(file_path)
    except Exception as e:
        logger.warning(f"Content validation failed for {file_path}: {e}")
        return False
    else:
        return True


def _validate_csv_content(file_path: str) -> bool:
    """Validate CSV file content.

    Args:
        file_path: Path to the CSV file

    Returns
    -------
        bool: True if content is valid
    """
    try:
        # Read only first few rows for validation
        df = pd.read_csv(file_path, nrows=10, encoding="utf-8")

        # Check for reasonable number of columns
        if len(df.columns) > FileSecurityConfig.MAX_COLUMNS:
            logger.warning(f"CSV file has too many columns: {len(df.columns)}")
            return False

        # Check for suspicious content in headers
        for col in df.columns:
            if _contains_suspicious_content(str(col)):
                logger.warning(f"Suspicious content in CSV header: {col}")
                return False

        # Check sample data for suspicious patterns
        for col in df.columns[:5]:  # Check first 5 columns
            sample_values = df[col].head(5).astype(str).tolist()
            for value in sample_values:
                if _contains_suspicious_content(value):
                    logger.warning(f"Suspicious content in CSV data: {value[:50]}")
                    return False

    except Exception as e:
        logger.warning(f"CSV validation error: {e}")
        return False
    else:
        return True


def _validate_excel_content(file_path: str) -> bool:
    """Validate Excel file content.

    Args:
        file_path: Path to the Excel file

    Returns
    -------
        bool: True if content is valid
    """
    try:
        from openpyxl import load_workbook

        # Load workbook in read-only mode
        wb = load_workbook(file_path, read_only=True, data_only=True)

        # Check number of sheets (prevent zip bombs)
        if len(wb.sheetnames) > 50:
            logger.warning(f"Excel file has too many sheets: {len(wb.sheetnames)}")
            wb.close()
            return False

        # Validate first sheet
        ws = wb.active
        if ws.max_column and ws.max_column > FileSecurityConfig.MAX_COLUMNS:
            logger.warning(f"Excel sheet has too many columns: {ws.max_column}")
            wb.close()
            return False

        # Check first few cells for suspicious content
        for row in range(1, min(6, ws.max_row + 1)):
            for col in range(1, min(6, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and _contains_suspicious_content(str(cell_value)):
                    logger.warning(f"Suspicious content in Excel cell: {cell_value}")
                    wb.close()
                    return False

        wb.close()
    except Exception as e:
        logger.warning(f"Excel validation error: {e}")
        return False
    else:
        return True


def _validate_json_content(file_path: str) -> bool:
    """Validate JSON file content.

    Args:
        file_path: Path to the JSON file

    Returns
    -------
        bool: True if content is valid
    """
    try:
        with open(file_path, encoding="utf-8") as f:
            # Limit JSON size during parsing
            content = f.read(10 * 1024 * 1024)  # 10MB limit
            data = json.loads(content)

        # Basic structure validation
        if isinstance(data, dict) and len(data) > 10000:
            logger.warning(f"JSON file has too many top-level keys: {len(data)}")
            return False

        # Check for suspicious content in string values
        def check_json_values(obj, depth=0):
            if depth > 10:  # Prevent deep recursion
                return True
            if isinstance(obj, str):
                return not _contains_suspicious_content(obj)
            elif isinstance(obj, dict):
                return all(
                    check_json_values(k, depth + 1) and check_json_values(v, depth + 1)
                    for k, v in list(obj.items())[:100]  # Limit items checked
                )
            elif isinstance(obj, list):
                return all(
                    check_json_values(item, depth + 1)
                    for item in obj[:100]  # Limit items checked
                )
            return True

        if not check_json_values(data):
            logger.warning("Suspicious content found in JSON file")
            return False

    except Exception as e:
        logger.warning(f"JSON validation error: {e}")
        return False
    else:
        return True


def _validate_stata_content(file_path: str) -> bool:
    """Validate Stata file content.

    Args:
        file_path: Path to the Stata file

    Returns
    -------
        bool: True if content is valid
    """
    try:
        # Basic file header validation
        with open(file_path, "rb") as f:
            header = f.read(100)

        # Check for basic Stata file patterns
        # Stata files typically start with specific byte patterns
        if len(header) < 4:
            return False

        # Basic size check - Stata files shouldn't be empty
        return not os.path.getsize(file_path) < 100
    except Exception as e:
        logger.warning(f"Stata validation error: {e}")
        return False


def _contains_suspicious_content(text: str) -> bool:
    """Check for suspicious content patterns.

    Args:
        text: Text content to check

    Returns
    -------
        bool: True if suspicious content is found
    """
    if not text or len(text) > 10000:  # Skip very long strings
        return False

    text_lower = text.lower()

    for pattern in FileSecurityConfig.SUSPICIOUS_PATTERNS:
        try:
            if re.search(pattern, text_lower, re.IGNORECASE | re.DOTALL):
                return True
        except re.error:
            continue  # Skip invalid regex patterns

    return False


def calculate_file_hash(file_path: str) -> str:
    """Calculate SHA-256 hash of file for integrity checking.

    Args:
        file_path: Path to the file

    Returns
    -------
        str: SHA-256 hash in hexadecimal
    """
    sha256_hash = hashlib.sha256()
    try:
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                sha256_hash.update(chunk)
    except Exception:
        logger.exception(f"Error calculating hash for {file_path}")
        return "error"
    return sha256_hash.hexdigest()


def get_file_info(file_path: str) -> dict[str, Any]:
    """Get comprehensive file information for security display.

    Args:
        file_path: Path to the file

    Returns
    -------
        Dict[str, Any]: File information dictionary
    """
    try:
        file_path_obj = Path(file_path)
        file_size = file_path_obj.stat().st_size
        file_hash = calculate_file_hash(file_path)

        return {
            "size_bytes": file_size,
            "size_mb": round(file_size / (1024 * 1024), 2),
            "extension": file_path_obj.suffix.upper(),
            "hash": file_hash,
            "hash_short": file_hash[:16] if file_hash != "error" else "error",
        }
    except Exception:
        logger.exception(f"Error getting file info for {file_path}")
        return {
            "size_bytes": 0,
            "size_mb": 0,
            "extension": "UNKNOWN",
            "hash": "error",
            "hash_short": "error",
        }


def _raise_unsupported_file_type(file_ext: str) -> None:
    """Raise SecurityError for unsupported file types."""
    raise SecurityError(f"Unsupported file type: {file_ext}")


def _raise_no_csv_data() -> None:
    """Raise SecurityError when no CSV data could be read."""
    raise SecurityError("No data could be read from CSV file")


def _raise_suspicious_content(col: str) -> None:
    """Raise SecurityError for suspicious content in column."""
    raise SecurityError(f"Suspicious content detected in column {col}")


def secure_read_data(filename: str, sheet_name: str | None = None) -> pl.DataFrame:
    """Secure version of data reading with enhanced validation.

    Args:
        filename: Path to the file to read
        sheet_name: Sheet name for Excel files (optional)

    Returns
    -------
        pl.DataFrame: Loaded data

    Raises
    ------
        SecurityError: If security validation fails
    """
    # Pre-read security validation
    is_valid, error_msg = validate_file_security(filename)
    if not is_valid:
        raise SecurityError(f"File security validation failed: {error_msg}")

    file_ext = Path(filename).suffix.lower().lstrip(".")

    try:
        # Use secure reading with limits
        if file_ext == "csv":
            data = _secure_read_csv(filename)
        elif file_ext in ["xlsx", "xls"]:
            data = _secure_read_excel(filename, sheet_name)
        elif file_ext == "json":
            data = _secure_read_json(filename)
        elif file_ext == "dta":
            data = _secure_read_stata(filename)
        else:
            _raise_unsupported_file_type(file_ext)

        # Post-read validation
        _validate_dataframe_security(data)

    except SecurityError:
        raise
    except Exception as e:
        logger.exception(f"Secure file read failed for {filename}")
        raise SecurityError(f"File reading failed: {e!s}") from e
    else:
        return data


def _secure_read_csv(filename: str) -> pl.DataFrame:
    """Secure CSV reading with size limits.

    Args:
        filename: Path to the CSV file

    Returns
    -------
        pl.DataFrame: Loaded data

    Raises
    ------
        SecurityError: If reading fails or file is too large
    """
    try:
        file_size = os.path.getsize(filename)

        # Read with chunk size for large files
        if file_size > 50 * 1024 * 1024:  # 50MB
            chunks = []
            chunk_size = 10000
            total_rows = 0

            for chunk in pd.read_csv(filename, chunksize=chunk_size, encoding="utf-8"):
                chunks.append(chunk)
                total_rows += len(chunk)
                if total_rows > 1000000:  # Limit total rows to 1M
                    logger.warning(f"CSV file truncated at {total_rows} rows")
                    break

            if chunks:
                data = pd.concat(chunks, ignore_index=True)
            else:
                _raise_no_csv_data()
        else:
            data = pd.read_csv(filename, encoding="utf-8")

        return pl.from_pandas(data)

    except Exception as e:
        raise SecurityError(f"CSV reading failed: {e!s}") from e


def _secure_read_excel(filename: str, sheet_name: str | None = None) -> pl.DataFrame:
    """Secure Excel reading with size limits.

    Args:
        filename: Path to the Excel file
        sheet_name: Sheet name to read

    Returns
    -------
        pl.DataFrame: Loaded data

    Raises
    ------
        SecurityError: If reading fails or file is too large
    """
    try:
        # Read with row limit for security
        data = pd.read_excel(
            filename,
            sheet_name=sheet_name,
            engine="openpyxl",
            nrows=1000000,  # Limit to 1M rows
        )

        return pl.from_pandas(data)

    except Exception as e:
        raise SecurityError(f"Excel reading failed: {e!s}") from e


def _check_json_size(filename: str) -> None:
    """Check if JSON file size is within limits.

    Args:
        filename: Path to the JSON file

    Raises
    ------
        SecurityError: If file is too large
    """
    file_size = os.path.getsize(filename)
    if file_size > 10 * 1024 * 1024:  # 10MB limit
        raise SecurityError("JSON file too large for secure reading")


def _secure_read_json(filename: str) -> pl.DataFrame:
    """Secure JSON reading with size limits.

    Args:
        filename: Path to the JSON file

    Returns
    -------
        pl.DataFrame: Loaded data

    Raises
    ------
        SecurityError: If reading fails or file is too large
    """
    try:
        _check_json_size(filename)
        data = pd.read_json(filename, encoding="utf-8")
        return pl.from_pandas(data)

    except Exception as e:
        raise SecurityError(f"JSON reading failed: {e!s}") from e


def _secure_read_stata(filename: str) -> pl.DataFrame:
    """Secure Stata file reading with size limits.

    Args:
        filename: Path to the Stata file

    Returns
    -------
        pl.DataFrame: Loaded data

    Raises
    ------
        SecurityError: If reading fails or file is too large
    """
    try:
        data = pd.read_stata(filename)
        return pl.from_pandas(data)

    except Exception as e:
        raise SecurityError(f"Stata reading failed: {e!s}") from e


def _validate_dataframe_security(df: pl.DataFrame) -> None:
    """Validate DataFrame doesn't contain malicious content.

    Args:
        df: DataFrame to validate

    Raises
    ------
        SecurityError: If DataFrame fails security validation
    """
    # Check DataFrame size
    if df.height > 1000000:  # 1M rows
        raise SecurityError("Dataset too large (>1M rows)")

    if df.width > FileSecurityConfig.MAX_COLUMNS:
        raise SecurityError(f"Too many columns (>{FileSecurityConfig.MAX_COLUMNS})")

    # Check for suspicious content in string columns
    string_cols = [col for col in df.columns if df[col].dtype == pl.Utf8]

    # Check first 10 string columns for performance
    for col in string_cols[:10]:
        try:
            sample_values = df[col].head(100).to_list()
            for value in sample_values:
                if value and _contains_suspicious_content(str(value)):
                    _raise_suspicious_content(col)
        except Exception as e:
            logger.warning(f"Error validating column {col}: {e}")
            continue


class VirusScanner:
    """Optional virus scanning interface."""

    @staticmethod
    def is_available() -> bool:
        """Check if virus scanning is available on the system.

        Returns
        -------
            bool: True if virus scanning is available
        """
        if os.name == "nt":  # Windows
            return VirusScanner._defender_available()
        else:  # Linux/macOS
            return VirusScanner._clamav_available()

    @staticmethod
    def _defender_available() -> bool:
        """Check if Windows Defender is available."""
        try:
            result = subprocess.run(
                ["powershell", "-Command", "Get-MpComputerStatus"],
                capture_output=True,
                text=True,
                timeout=10,
            )
        except Exception:
            return False
        else:
            return result.returncode == 0

    @staticmethod
    def _clamav_available() -> bool:
        """Check if ClamAV is available."""
        try:
            result = subprocess.run(
                ["clamdscan", "--version"], capture_output=True, text=True, timeout=10
            )
        except Exception:
            return False
        else:
            return result.returncode == 0

    @staticmethod
    def scan_file(file_path: str) -> tuple[bool, str]:
        """Scan file for viruses.

        Args:
            file_path: Path to the file to scan

        Returns
        -------
            Tuple[bool, str]: (is_clean, result_message)
        """
        try:
            if os.name == "nt":
                return VirusScanner._scan_with_defender(file_path)
            else:
                return VirusScanner._scan_with_clamav(file_path)
        except Exception as e:
            logger.warning(f"Virus scan failed for {file_path}: {e!s}")
            return True, "Scan unavailable"

    @staticmethod
    def _scan_with_defender(file_path: str) -> tuple[bool, str]:
        """Scan using Windows Defender."""
        try:
            result = subprocess.run(
                [
                    "powershell",
                    "-Command",
                    f'Start-MpScan -ScanType CustomScan -ScanPath "{file_path}"',
                ],
                capture_output=True,
                text=True,
                timeout=30,
            )

            if result.returncode == 0:
                return True, "Clean"
            else:
                return False, "Threat detected"
        except Exception:
            return True, "Scan failed"

    @staticmethod
    def _scan_with_clamav(file_path: str) -> tuple[bool, str]:
        """Scan using ClamAV."""
        try:
            result = subprocess.run(
                ["clamdscan", "--no-summary", file_path],
                capture_output=True,
                text=True,
                timeout=30,
            )

            if "OK" in result.stdout:
                return True, "Clean"
            else:
                return False, "Threat detected"
        except Exception:
            return True, "Scan failed"
