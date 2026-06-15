"""Tests for the local connector module."""

import json
import os
from unittest.mock import MagicMock, patch

import polars as pl
import pytest
from openpyxl import Workbook
from pydantic import ValidationError

from datasure.connectors.local import (
    VALID_FILE_TYPES,
    FileConfig,
    ImportLogEntry,
    get_excel_sheet_names,
    get_file_info,
    load_data_efficiently,
    load_local_data,
    render_local_file_form,
    validate_file_accessibility,
)


@pytest.fixture(autouse=True)
def mock_database_functions(monkeypatch):
    """Override the autouse fixture from conftest.

    Disables database mocking for these tests.
    """
    pass


class TestValidFileTypes:
    """Test the VALID_FILE_TYPES constant."""

    def test_valid_file_types_constant(self):
        """Test that VALID_FILE_TYPES contains expected file extensions."""
        expected_types = {".csv", ".xlsx", ".xls", ".json", ".dta"}
        assert expected_types == VALID_FILE_TYPES


class TestFileConfig:
    """Test the FileConfig Pydantic model."""

    def test_valid_file_config(self, tmp_path):
        """Test creating a valid FileConfig."""
        # Create a test file
        test_file = tmp_path / "test.csv"
        test_file.write_text("name,age\nJohn,25")

        config = FileConfig(
            alias="test_alias", filename=str(test_file), sheet_name="Sheet1"
        )

        assert config.alias == "test_alias"
        assert config.filename == str(test_file)
        assert config.sheet_name == "Sheet1"
        assert config.source == "local storage"

    def test_alias_validation_empty(self):
        """Test alias validation with empty alias."""
        with pytest.raises(ValidationError) as exc_info:
            FileConfig(alias="", filename="/path/to/file.csv")

        assert "String should have at least 1 character" in str(exc_info.value)

    def test_alias_validation_whitespace_only(self):
        """Test alias validation with whitespace-only alias."""
        with pytest.raises(ValidationError) as exc_info:
            FileConfig(alias="   ", filename="/path/to/file.csv")

        assert "Alias cannot be empty" in str(exc_info.value)

    def test_alias_validation_too_long(self):
        """Test alias validation with too long alias."""
        long_alias = "a" * 21  # 21 characters
        with pytest.raises(ValidationError) as exc_info:
            FileConfig(alias=long_alias, filename="/path/to/file.csv")

        assert "at most 20 characters" in str(exc_info.value)

    def test_alias_validation_strips_whitespace(self, tmp_path):
        """Test that alias validation strips whitespace."""
        test_file = tmp_path / "test.csv"
        test_file.write_text("test")

        config = FileConfig(alias="  test_alias  ", filename=str(test_file))
        assert config.alias == "test_alias"

    def test_filename_validation_empty(self):
        """Test filename validation with empty filename."""
        with pytest.raises(ValidationError) as exc_info:
            FileConfig(alias="test", filename="")

        assert "File path cannot be empty" in str(exc_info.value)

    def test_filename_validation_nonexistent_file(self):
        """Test filename validation with non-existent file."""
        with pytest.raises(ValidationError) as exc_info:
            FileConfig(alias="test", filename="/nonexistent/file.csv")

        assert "File not found" in str(exc_info.value)

    def test_filename_validation_directory(self, tmp_path):
        """Test filename validation with directory path."""
        with pytest.raises(ValidationError) as exc_info:
            FileConfig(alias="test", filename=str(tmp_path))

        assert "Path is not a file" in str(exc_info.value)

    def test_filename_validation_invalid_extension(self, tmp_path):
        """Test filename validation with invalid file extension."""
        test_file = tmp_path / "test.txt"
        test_file.write_text("test")

        with pytest.raises(ValidationError) as exc_info:
            FileConfig(alias="test", filename=str(test_file))

        assert "Invalid file type" in str(exc_info.value)
        assert ".csv" in str(exc_info.value)


class TestImportLogEntry:
    """Test the ImportLogEntry Pydantic model."""

    def test_import_log_entry_defaults(self):
        """Test ImportLogEntry with minimal required fields."""
        entry = ImportLogEntry(alias="test", filename="test.csv")

        assert entry.refresh is True
        assert entry.load is True
        assert entry.alias == "test"
        assert entry.filename == "test.csv"
        assert entry.sheet_name is None
        assert entry.source == "local storage"
        assert entry.server == ""
        assert entry.username == ""

    def test_import_log_entry_all_fields(self):
        """Test ImportLogEntry with all fields."""
        entry = ImportLogEntry(
            refresh=False,
            load=False,
            alias="test",
            filename="test.csv",
            sheet_name="Sheet1",
            source="custom source",
            server="test_server",
            username="test_user",
            form_id="form123",
            private_key="key123",
            save_to="location",
            attachments=True,
        )

        assert entry.refresh is False
        assert entry.load is False
        assert entry.sheet_name == "Sheet1"
        assert entry.source == "custom source"


class TestGetExcelSheetNames:
    """Test the get_excel_sheet_names function."""

    def test_excel_sheet_names_single_sheet(self, tmp_path):
        """Test getting sheet names from Excel file with single sheet."""
        excel_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "test data"
        wb.save(excel_file)

        result = get_excel_sheet_names(str(excel_file))
        assert result == ["TestSheet"]

    def test_excel_sheet_names_multiple_sheets(self, tmp_path):
        """Test getting sheet names from Excel file with multiple sheets."""
        excel_file = tmp_path / "test.xlsx"
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Sheet1"
        wb.create_sheet("Sheet2")
        wb.create_sheet("Sheet3")

        wb.save(excel_file)

        result = get_excel_sheet_names(str(excel_file))
        assert result == ["Sheet1", "Sheet2", "Sheet3"]

    @patch("datasure.connectors.local.st")
    @patch("datasure.connectors.local.load_workbook")
    def test_excel_sheet_names_error_handling(self, mock_load_workbook, mock_st):
        """Test error handling in get_excel_sheet_names."""
        mock_load_workbook.side_effect = OSError("Cannot read file")

        result = get_excel_sheet_names("invalid_file.xlsx")

        assert result == []
        mock_st.error.assert_called_once()
        assert "Error reading Excel file" in mock_st.error.call_args[0][0]


class TestLoadDataEfficiently:
    """Test the load_data_efficiently function."""

    def test_load_csv_file(self, tmp_path):
        """Test loading CSV file."""
        csv_file = tmp_path / "test.csv"
        test_data = "name,age,city\nJohn,25,NYC\nJane,30,LA"
        csv_file.write_text(test_data)

        result = load_data_efficiently(str(csv_file))

        assert isinstance(result, pl.DataFrame)
        assert result.shape[0] == 2
        assert list(result.columns) == ["name", "age", "city"]
        assert result.row(0) == ("John", 25, "NYC")

    def test_load_excel_file_default_sheet(self, tmp_path):
        """Test loading Excel file with default sheet."""
        excel_file = tmp_path / "test.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "name"
        ws["B1"] = "age"
        ws["A2"] = "John"
        ws["B2"] = 25
        wb.save(excel_file)

        result = load_data_efficiently(filename=excel_file, sheet_name="TestSheet")

        assert isinstance(result, pl.DataFrame)
        assert result.shape[0] == 1
        assert "name" in result.columns
        assert "age" in result.columns

    def test_load_excel_file_specific_sheet(self, tmp_path):
        """Test loading Excel file with specific sheet name."""
        excel_file = tmp_path / "test.xlsx"

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "data1"

        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "data2"
        ws2["A2"] = "value2"

        wb.save(excel_file)

        result = load_data_efficiently(str(excel_file), sheet_name="Sheet2")

        assert isinstance(result, pl.DataFrame)
        assert result.shape[0] >= 1

    def test_load_json_file(self, tmp_path):
        """Test loading JSON file."""
        json_file = tmp_path / "test.json"
        test_data = [
            {"name": "John", "age": 25, "city": "NYC"},
            {"name": "Jane", "age": 30, "city": "LA"},
        ]
        json_file.write_text(json.dumps(test_data))

        result = load_data_efficiently(str(json_file))

        assert isinstance(result, pl.DataFrame)
        assert result.shape[0] == 2
        assert "name" in result.columns

    @patch("datasure.connectors.local.scan_readstat")
    def test_load_stata_file(self, mock_scan_readstat, tmp_path):
        """Test loading Stata file."""
        stata_file = tmp_path / "test.dta"
        stata_file.touch()

        mock_lazy_frame = MagicMock()
        mock_df = pl.DataFrame({"var1": [1, 2, 3], "var2": ["a", "b", "c"]})
        mock_lazy_frame.collect.return_value = mock_df
        mock_scan_readstat.return_value = mock_lazy_frame

        result = load_data_efficiently(str(stata_file))

        mock_scan_readstat.assert_called_once_with(str(stata_file))
        assert result.equals(mock_df)

    def test_unsupported_file_format(self, tmp_path):
        """Test loading unsupported file format."""
        unsupported_file = tmp_path / "test.txt"
        unsupported_file.write_text("test")

        with patch("datasure.connectors.local.st") as mock_st:
            result = load_data_efficiently(str(unsupported_file))

            assert isinstance(result, pl.DataFrame)
            assert result.is_empty()
            mock_st.error.assert_called_once()
            assert "Error loading file" in mock_st.error.call_args[0][0]

    @patch("datasure.connectors.local.pl.read_csv")
    @patch("datasure.connectors.local.st")
    def test_load_data_error_handling(self, mock_st, mock_read_csv, tmp_path):
        """Test error handling in load_data_efficiently."""
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("test")

        mock_read_csv.side_effect = pl.exceptions.ComputeError("Read error")

        result = load_data_efficiently(str(csv_file))

        assert isinstance(result, pl.DataFrame)
        assert result.is_empty()
        mock_st.error.assert_called_once()
        assert "Error loading file" in mock_st.error.call_args[0][0]


class TestRenderLocalFileForm:
    """Test the render_local_file_form function."""

    @patch("datasure.connectors.local.st")
    @patch("datasure.connectors.local.Path")
    def test_render_form_basic(self, mock_path, mock_st):
        """Test basic form rendering."""
        # Mock Path and file system
        mock_assets_dir = MagicMock()
        mock_image_path = MagicMock()
        mock_image_path.exists.return_value = True
        mock_assets_dir.__truediv__.return_value = mock_image_path
        mock_path.return_value.parent.parent.__truediv__.return_value = mock_assets_dir

        # Mock form context
        mock_form = MagicMock()
        mock_st.form.return_value.__enter__.return_value = mock_form

        # Mock form inputs
        mock_st.text_input.side_effect = ["test_alias", "/path/to/file.csv"]
        mock_st.form_submit_button.return_value = False

        render_local_file_form("test_project_id")

        # Verify UI elements were called
        mock_st.image.assert_called_once()
        mock_st.subheader.assert_called_with("Add File from Local Storage")
        assert mock_st.text_input.call_count == 2

    @patch("datasure.connectors.local.st")
    @patch("datasure.connectors.local.Path")
    def test_render_form_edit_mode(self, mock_path, mock_st):
        """Test form rendering in edit mode."""
        # Mock Path setup
        mock_assets_dir = MagicMock()
        mock_image_path = MagicMock()
        mock_image_path.exists.return_value = False
        mock_assets_dir.__truediv__.return_value = mock_image_path
        mock_path.return_value.parent.parent.__truediv__.return_value = mock_assets_dir

        # Mock form context
        mock_form = MagicMock()
        mock_st.form.return_value.__enter__.return_value = mock_form

        defaults = {"alias": "old_alias", "filename": "old_file.csv"}
        mock_st.text_input.side_effect = ["old_alias", "old_file.csv"]
        mock_st.form_submit_button.return_value = False

        render_local_file_form("test_project_id", edit_mode=True, defaults=defaults)

        # Verify edit mode info message
        mock_st.info.assert_called_with(
            "You are in edit mode. Please modify the file details below."
        )

        # Verify alias input is disabled in edit mode
        alias_call = mock_st.text_input.call_args_list[0]
        assert alias_call[1]["disabled"] is True
        assert alias_call[1]["value"] == "old_alias"

    @patch("datasure.connectors.local._handle_form_submission")
    @patch("datasure.connectors.local.get_excel_sheet_names")
    @patch("datasure.connectors.local.st")
    @patch("datasure.connectors.local.Path")
    def test_render_form_excel_sheet_selection(
        self, mock_path, mock_st, mock_get_sheets, mock_handle_submission
    ):
        """Test Excel sheet selection in form."""
        # Mock Path setup for Excel file
        mock_path_obj = MagicMock()
        mock_path_obj.suffix.lower.return_value = ".xlsx"
        mock_path_obj.exists.return_value = True
        mock_path.return_value = mock_path_obj

        # Mock assets path
        mock_assets_dir = MagicMock()
        mock_image_path = MagicMock()
        mock_image_path.exists.return_value = True
        mock_assets_dir.__truediv__.return_value = mock_image_path
        mock_path.return_value.parent.parent.__truediv__.return_value = mock_assets_dir

        # Mock form context
        mock_form = MagicMock()
        mock_st.form.return_value.__enter__.return_value = mock_form

        # Mock Excel sheets
        mock_get_sheets.return_value = ["Sheet1", "Sheet2", "Data"]

        # Mock form inputs
        mock_st.text_input.side_effect = ["test_alias", "/path/to/file.xlsx"]
        mock_st.selectbox.return_value = "Sheet2"
        mock_st.form_submit_button.return_value = True

        render_local_file_form("test_project_id")

        # Verify form submission was handled
        mock_handle_submission.assert_called_once()


class TestHandleFormSubmission:
    """Test the _handle_form_submission function (indirectly through form rendering)."""

    @patch("datasure.connectors.local.duckdb_save_table")
    @patch("datasure.connectors.local.duckdb_get_table")
    @patch("datasure.connectors.local.st")
    @patch("datasure.connectors.local.Path")
    def test_successful_form_submission(
        self, mock_path, mock_st, mock_get_table, mock_save_table
    ):
        """Test successful form submission."""
        # Create a real temp file for validation
        import tempfile

        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tmp:
            tmp.write(b"test,data\n1,2")
            temp_file_path = tmp.name

        try:
            # Mock Path for assets
            mock_assets_dir = MagicMock()
            mock_image_path = MagicMock()
            mock_image_path.exists.return_value = True
            mock_assets_dir.__truediv__.return_value = mock_image_path
            mock_path.return_value.parent.parent.__truediv__.return_value = (
                mock_assets_dir
            )

            # Mock form context
            mock_form = MagicMock()
            mock_st.form.return_value.__enter__.return_value = mock_form

            # Mock empty import log
            mock_import_log = MagicMock()
            mock_import_log.is_empty.return_value = True
            mock_get_table.return_value = mock_import_log

            # Mock form inputs - using real file path
            mock_st.text_input.side_effect = ["test_alias", temp_file_path]
            mock_st.form_submit_button.return_value = True

        finally:
            # Clean up temp file
            os.unlink(temp_file_path)


class TestLoadLocalData:
    """Test the load_local_data function."""

    @patch("datasure.connectors.local.duckdb_save_table")
    @patch("datasure.connectors.local.load_data_efficiently")
    @patch("datasure.connectors.local.st")
    def test_load_local_data_success(self, mock_st, mock_load_data, mock_save_table):
        """Test successful data loading."""
        # Mock data loading
        mock_df = pl.DataFrame({"col1": [1, 2, 3], "col2": ["a", "b", "c"]})
        mock_load_data.return_value = mock_df

        load_local_data("test_project", "test_alias", "/path/to/file.csv")

        # Verify functions were called correctly
        mock_load_data.assert_called_once_with("/path/to/file.csv", None)
        mock_save_table.assert_called_once_with(
            "test_project", mock_df, alias="test_alias", db_name="raw"
        )
        mock_st.success.assert_called_once()
        assert "Shape: (3, 2)" in mock_st.success.call_args[0][0]

    @patch("datasure.connectors.local.duckdb_save_table")
    @patch("datasure.connectors.local.load_data_efficiently")
    @patch("datasure.connectors.local.st")
    def test_load_local_data_with_sheet(self, mock_st, mock_load_data, mock_save_table):
        """Test data loading with Excel sheet name."""
        mock_df = pl.DataFrame({"col1": [1, 2], "col2": ["x", "y"]})
        mock_load_data.return_value = mock_df

        load_local_data("test_project", "test_alias", "/path/to/file.xlsx", "Sheet2")

        mock_load_data.assert_called_once_with("/path/to/file.xlsx", "Sheet2")
        mock_save_table.assert_called_once()

    @patch("datasure.connectors.local.load_data_efficiently")
    @patch("datasure.connectors.local.st")
    def test_load_local_data_empty_result(self, mock_st, mock_load_data):
        """Test data loading with empty result."""
        mock_df = pl.DataFrame()
        mock_load_data.return_value = mock_df

        load_local_data("test_project", "test_alias", "/path/to/file.csv")

        mock_st.warning.assert_called_once()
        assert "No data loaded" in mock_st.warning.call_args[0][0]

    @patch("datasure.connectors.local.load_data_efficiently")
    @patch("datasure.connectors.local.st")
    def test_load_local_data_error(self, mock_st, mock_load_data):
        """Test data loading error handling."""
        mock_load_data.side_effect = Exception("Load error")

        load_local_data("test_project", "test_alias", "/path/to/file.csv")

        mock_st.error.assert_called_once()
        assert "Error loading data" in mock_st.error.call_args[0][0]


class TestValidateFileAccessibility:
    """Test the validate_file_accessibility function."""

    def test_validate_accessible_file(self, tmp_path):
        """Test validation of accessible file."""
        test_file = tmp_path / "test.txt"
        test_file.write_text("test content")

        result = validate_file_accessibility(str(test_file))
        assert result is True

    def test_validate_nonexistent_file(self):
        """Test validation of non-existent file."""
        result = validate_file_accessibility("/nonexistent/file.txt")
        assert result is False

    def test_validate_directory_path(self, tmp_path):
        """Test validation of directory path."""
        result = validate_file_accessibility(str(tmp_path))
        assert result is False

    @patch("datasure.connectors.local.os.access")
    def test_validate_unreadable_file(self, mock_access, tmp_path):
        """Test validation of unreadable file."""
        test_file = tmp_path / "test.txt"
        test_file.write_text("test")

        mock_access.return_value = False

        result = validate_file_accessibility(str(test_file))
        assert result is False


class TestGetFileInfo:
    """Test the get_file_info function."""

    def test_get_info_existing_file(self, tmp_path):
        """Test getting info for existing file."""
        test_file = tmp_path / "test.csv"
        test_file.write_text("name,age\nJohn,25")

        info = get_file_info(str(test_file))

        assert info["exists"] is True
        assert info["size"] > 0
        assert info["extension"] == ".csv"
        assert "modified" in info

    def test_get_info_nonexistent_file(self):
        """Test getting info for non-existent file."""
        info = get_file_info("/nonexistent/file.txt")

        assert info["exists"] is False
        assert len(info) == 1  # Only contains 'exists' key

    @patch("datasure.connectors.local.Path")
    def test_get_info_error_handling(self, mock_path):
        """Test error handling in get_file_info."""
        mock_path.side_effect = OSError("Path error")

        info = get_file_info("some_path")

        assert info["exists"] is False
